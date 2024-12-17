import os
from flask import Blueprint, json, render_template, request, redirect, session, jsonify, url_for, current_app, flash, make_response, send_file
import openpyxl
from website.models import db, Product, Order, Purchase, Supplier
from website.functions import login_required
from datetime import datetime
from fpdf import FPDF
import io
from werkzeug.utils import secure_filename
from num2words import num2words



views = Blueprint('views', __name__)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'heic', 'heif'}

@views.route("/")
@login_required
def index():
    try:
        totals = db.session.query(
            db.func.count(Product.productName).label("total_products"),
            db.func.sum(Product.inventoryOut).label("total_orders"),
            db.func.sum(Product.inventoryIn).label("total_purchases")
        ).first()

        ups = db.session.query(
            db.func.count(Product.id).label("total_ups")
        ).filter(Product.inventoryOnHand < Product.minimumReq, Product.minimumReq != 0).first()

        return render_template("index.html", totals=totals, ups=ups)

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect("/")

@views.route("/get_data")
def get_data():
    try:
        # Total counts
        totals = db.session.query(
            db.func.count(Product.productName).label("total_products"),
            db.func.sum(Product.inventoryOut).label("total_orders"),
            db.func.sum(Product.inventoryIn).label("total_purchases")
        ).first()

        labels = ["Total Products", "Total Orders", "Total Purchases"]
        data = [totals.total_products or 0, totals.total_orders or 0, totals.total_purchases or 0]

        # Monthly totals
        mtotal = db.session.query(
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '01', 1), else_=0)).label('jan'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '02', 1), else_=0)).label('feb'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '03', 1), else_=0)).label('mar'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '04', 1), else_=0)).label('apr'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '05', 1), else_=0)).label('may'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '06', 1), else_=0)).label('jun'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '07', 1), else_=0)).label('jul'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '08', 1), else_=0)).label('aug'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '09', 1), else_=0)).label('sep'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '10', 1), else_=0)).label('oct'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '11', 1), else_=0)).label('nov'),
            db.func.sum(db.case((db.func.strftime('%m', Order.oDate) == '12', 1), else_=0)).label('dec')
        ).first()

        mtotals = [getattr(mtotal, month.lower(), 0) for month in ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']]

        # Return as JSON response
        return jsonify({
            'payload': json.dumps({'data': data, 'labels': labels, 'mtotals': mtotals})
        })

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect("/")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@views.route("/productv", methods=["GET", "POST"])
@login_required
def productv():
    try:
        if request.method == "POST":
            product_name = request.form.get("product-name")
            part_number = request.form.get("part-number")

            # Check if the product already exists
            existing_product = Product.query.filter_by(
                productName=product_name,
                partNumber=part_number
            ).first()

            if existing_product:
                flash('Product with the same name and part number already exists.', 'error')
                return redirect(request.url)

            # Handle file upload
            file = request.files.get('product-image')  # Use .get() to safely check if file is present

            if file and allowed_file(file.filename):  # Valid file
                filename = secure_filename(f"{product_name}_{part_number}.{file.filename.rsplit('.', 1)[1].lower()}")
                filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                image_path = url_for('static', filename=f'img/{filename}')
            else:  # No file or invalid file type
                image_path = url_for('static', filename='img/empty.jpg')

            # Create a new product
            new_product = Product(
                productName=product_name,
                partNumber=part_number,
                productLabel=request.form.get("product-label"),
                startingInventory=request.form.get("starting-inventory"),
                inventoryIn=request.form.get("inventory-in"),
                inventoryOut=request.form.get("inventory-out"),
                inventoryOnHand=request.form.get("inventory-on-hand"),
                minimumReq=request.form.get("minimum-req"),
                image=image_path,  # Use the path determined above
                sale_price=request.form.get("sale-price"),
                buy_price=request.form.get("buy-price")
            )
            db.session.add(new_product)
            db.session.commit()
            flash('Product added successfully.', 'success')
            return redirect("/productv")

        return redirect("/productl")

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect("/")


@views.route("/purchasev", methods=["GET", "POST"])
@login_required
def purchasev():
    try:
        if request.method == "POST":
            pDate_str = request.form.get("purchase-date")
            pDate = datetime.strptime(pDate_str, '%Y-%m-%d').date()
            purchases = int(request.form.get("purchases"))
            product_name, part_number = request.form.get("product-name").split("|")
            if purchases <= 0:
                return "Purchases must be greater than 0"
            
            product = db.session.query(Product).filter(
                Product.productName == product_name,
                Product.partNumber == part_number
            ).first()

            if not product:
                flash("Invalid product selection.", "error")
                return redirect("/purchasev")
            
            product.inventoryOnHand += purchases
            product.inventoryIn += purchases
            db.session.add(product)

            purchase_price = product.buy_price * purchases
            
            new_purchase = Purchase(
                supplier_name=request.form.get("supplier-name"),
                product_name=product_name,
                part_number=part_number,
                numberIn=purchases,
                pDate=pDate,
                purchase_price=purchase_price
            )
            db.session.add(new_purchase)
            db.session.commit()
            
            flash('Purchase added successfully.', 'success')
            return redirect("/purchasel")
        
        products = db.session.query(Product.id, Product.productName, Product.partNumber).all()
        suppliers = db.session.query(Supplier.id, Supplier.supplier).all()
        return render_template("purchasev.html", products=products, suppliers=suppliers)

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect("/purchasev")

@views.route("/supplierv", methods=["GET", "POST"])
@login_required
def supplierv():
    try:
        if request.method == "POST":
            new_supplier = Supplier(
                supplier=request.form.get("supplier-name")
            )
            db.session.add(new_supplier)
            db.session.commit()
            flash('Supplier added successfully.', 'success')
            return redirect("/supplierl")
        
        return render_template("supplierv.html")

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect("/supplierv")

@views.route("/productl")
@login_required
def productl():
    try:
        products = db.session.query(Product).all()
        return render_template("productl.html", rows=products)

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect("/productl")

@views.route("/orderl")
@login_required
def orderl():
    print("Orderl")
    try:
        # Fetch all orders with their associated user, order_items, and products
        orders = Order.query.options(
            db.joinedload(Order.user),
            db.joinedload(Order.order_items).joinedload(OrderItem.product)
        ).all()

        # Prepare the order data
        order_data = []
        for order in orders:
            order_items = []
            total_order_price = 0
            for item in order.order_items:
                total_item_price = item.product.sale_price * item.quantity
                order_items.append({
                    "product_name": item.product.productName,
                    "part_number": item.product.partNumber,
                    "quantity": item.quantity,
                    "price": item.product.sale_price,
                    "total_price": total_item_price
                })
                total_order_price += total_item_price

            order_data.append({
                "orderId": order.orderId,
                "user": order.user.fullname,
                "order_items": order_items,
                "total_order_price": total_order_price,
                "oDate": order.oDate,
                "cust_name": order.cust_name,
                "cust_phone": order.cust_phone
            })
        
        return render_template("orderl.html", orders=order_data)

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        print(e)
        return redirect("/orderl")

@views.route("/purchasel")
@login_required
def purchasel():
    try:
        purchases = db.session.query(Purchase).all()
        products = db.session.query(Product.id, Product.productName, Product.partNumber).all()
        suppliers = db.session.query(Supplier).all()
        return render_template("purchasel.html", rows=purchases, products=products, suppliers=suppliers)

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect("/purchasel")

@views.route("/supplierl")
@login_required
def supplierl():
    try:
        suppliers = db.session.query(Supplier).all()
        return render_template("supplierl.html", suppliers=suppliers)

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect("/supplierl")

@views.route("/up")
@login_required
def up():
    try:
        products = db.session.query(Product).filter(Product.inventoryOnHand < Product.minimumReq, Product.minimumReq != 0).all()
        return render_template("up.html", products=products)

    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect("/up")


from flask import render_template, request, redirect, url_for, flash, session
from .models import OrderItem, db, Product, Cart, CartItem, Order

@views.route('/shop', methods=['GET', 'POST'])
def shop():
    search_query = request.args.get('search', '')
    if search_query:
        products = Product.query.filter(Product.productName.contains(search_query)).all()
    else:
        products = Product.query.all()
    return render_template('shop.html', products=products)


@views.route('/add_to_cart', methods=['POST'])
def add_to_cart():
    product_id = request.form.get('product_id')
    quantity = int(request.form.get('quantity', 1))
    
    user_id = session.get('user_id')
    cart = Cart.query.filter_by(user_id=user_id).first()
    if not cart:
        cart = Cart(user_id=user_id)
        db.session.add(cart)
        db.session.commit()
    
    cart_item = CartItem.query.filter_by(cart_id=cart.id, product_id=product_id).first()
    if cart_item:
        cart_item.quantity += quantity
    else:
        cart_item = CartItem(cart_id=cart.id, product_id=product_id, quantity=quantity)
        db.session.add(cart_item)
    
    db.session.commit()
    flash("Product added to cart!", "success")
    return redirect("/shop")

@views.route('/checkout', methods=['POST'])
def checkout():
    user_id = session.get('user_id')
    
    # Get the user's cart
    cart = Cart.query.filter_by(user_id=user_id).first()
    
    if not cart or not cart.cart_items:
        flash("Cart is empty!", "warning")
        return redirect(url_for('views.shop'))

    # Calculate the total amount for the order
    total_amount = 0
    for item in cart.cart_items:
        total_amount += item.product.sale_price * item.quantity  # Replace with actual price logic

    if total_amount == 0:
        flash("Purchase Failed!", "warning")
        return redirect(url_for('views.shop'))
    # Create the order record
    order = Order(
        user_id=user_id,
        oDate=datetime.now(),
        totalAmount=total_amount,
        cust_phone=request.form.get('cust_phone'),
        cust_name=request.form.get('cust_name')
    )
    
    db.session.add(order)
    db.session.commit()  # Commit to get the order ID

    # Create the order items and update product inventory
    for item in cart.cart_items:
        product = Product.query.get(item.product.id)  # Get the product
        
        # Check if enough inventory is available
        if product.inventoryOnHand < item.quantity:
            flash(f"Not enough inventory for {product.productName}.", "danger")
            return redirect(url_for('views.cart'))  # Redirect to the cart if inventory is insufficient

        # Deduct the ordered quantity from inventory
        product.inventoryOut += item.quantity
        product.inventoryOnHand -= item.quantity

        # Create the order item
        order_item = OrderItem(
            order_id=order.orderId,
            product_id=item.product.id,
            quantity=item.quantity,
        )
        db.session.add(order_item)
        db.session.add(product)  # Save updated product inventory

    # Commit all the changes
    db.session.commit()

    # Clear the cart
    db.session.delete(cart)
    db.session.commit()

    flash("Order placed successfully!", "success")
    return redirect(url_for('views.shop'))



@views.route('/cart')
def cart():
    user_id = session.get('user_id')  # Get user_id from session or another method
    cart = Cart.query.filter_by(user_id=user_id).first()
    
    if not cart:
        flash('Your cart is empty.', 'info')
        return render_template('cart.html', cart=None)

    return render_template('cart.html', cart=cart)

@views.route('/remove_from_cart/<int:item_id>', methods=['POST'])
def remove_from_cart(item_id):
    cart_item = CartItem.query.get(item_id)
    if cart_item:
        db.session.delete(cart_item)
        db.session.commit()
        flash('Item removed from cart!', 'success')
    else:
        flash('Item not found in cart!', 'danger')
    
    return redirect(url_for('views.cart'))

@views.route('/update_cart_item/<int:item_id>', methods=['POST'])
def update_cart_item(item_id):
    new_quantity = int(request.form['quantity'])
    cart_item = CartItem.query.get(item_id)
    
    if cart_item:
        if new_quantity > 0:
            cart_item.quantity = new_quantity
            db.session.commit()
            flash('Cart updated!', 'success')
        else:
            flash('Invalid quantity!', 'danger')
    else:
        flash('Item not found in cart!', 'danger')
    
    return redirect(url_for('views.cart'))


class PDF(FPDF):
    def header(self):
        self.set_font("Arial", size=12)
        self.cell(0, 10, "Mani Electricals", align="C", ln=1)

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", size=10)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")

class PDF(FPDF):
    def header(self):
        self.set_font("Arial", size=12)
        self.cell(0, 10, "Mani Electricals", align="C", ln=1)

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", size=10)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")

def generate_bill(order, order_items):
    pdf = PDF()
    pdf.add_page()

    # Header
    pdf.set_font("Arial", size=12)
    pdf.cell(95, 10, f"Order ID: {order.orderId}", ln=0,align="C")
    pdf.cell(0, 10, f"Sale Date: {order.oDate.strftime('%d-%m-%Y')}", ln=1, align="C")

    # Customer Details
    pdf.cell(0, 10, f"Customer Name: {order.cust_name} | Customer Phone: {order.cust_phone}", ln=1, align="C")
    pdf.cell(0, 10, f"Billed By: {order.user.fullname if order.user else 'Unknown'}", ln=1,align="C")
    pdf.ln(10)  # Line break

    # Table Headers
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(10, 10, "S.No", border=1, align="C")
    pdf.cell(60, 10, "Name-Part Number/Code", border=1, align="C")
    pdf.cell(30, 10, "Price (Rs.)", border=1, align="C")
    pdf.cell(20, 10, "Qty", border=1, align="C")
    pdf.cell(40, 10, "Total Price (Rs.)", border=1, align="C")
    pdf.ln(10)

    # Table Content
    pdf.set_font("Arial", size=12)
    for i, item in enumerate(order_items, start=1):
        product = item.product
        pdf.cell(10, 10, str(i), border=1, align="C")
        pdf.cell(60, 10, f"{product.productName} - {product.partNumber}", border=1)
        pdf.cell(30, 10, f"Rs.{product.sale_price:.2f}", border=1, align="C")
        pdf.cell(20, 10, str(item.quantity), border=1, align="C")
        pdf.cell(40, 10, f"Rs.{product.sale_price * item.quantity:.2f}", border=1, align="R")
        pdf.ln(10)

    # Total Price4
    total_price = order.totalAmount
    total_price_in_words = num2words(total_price)

    pdf.ln(10)
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(0, 10, f"Total Price: Rs.{order.totalAmount:.2f}")
    pdf.ln(5)
    pdf.cell(0, 10,f"(Rupees {total_price_in_words.capitalize()} only)")

    pdf.ln(40)
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(50, 10, f"Authorized Signatory")

    # Return the PDF as a bytes object
    pdf_output = io.BytesIO()
    pdf.output(pdf_output)
    pdf_output.seek(0)
    return pdf_output


@views.route('/download-bill/<int:order_id>', methods=['GET'])
@login_required
def download_bill(order_id):
    # Fetch the order and associated items
    order = Order.query.get(order_id)
    if not order:
        return jsonify({"error": "Order not found"}), 404
    
    order_items = OrderItem.query.filter_by(order_id=order_id).all()
    if not order_items:
        return jsonify({"error": "No items found for this order"}), 404

    # Generate the PDF and return it
    pdf_data = generate_bill(order, order_items)
    return send_file(pdf_data, download_name=f"order_{order_id}.pdf", as_attachment=True)

@views.route("/upload-products", methods=["POST"])
@login_required
def upload_products():
    try:
        file = request.files.get('product-xlsx')  # Expecting a .xlsx file from the form

        # Check if a file is provided and if it is of the correct type
        if not file or not file.filename.endswith('.xlsx'):
            flash("Please upload a valid .xlsx file.", 'error')
            return redirect(request.url)

        # Secure the filename and save it temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Load the workbook and the active sheet
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # Iterate over the rows of the sheet (assuming headers are in the first row)
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            product_name, part_number, product_label, starting_inventory, inventory_in, inventory_out, inventory_on_hand, minimum_req, sale_price, buy_price = row

            # Check if the product already exists in the database
            existing_product = Product.query.filter_by(
                productName=product_name,
                partNumber=part_number
            ).first()

            if existing_product:
                flash(f"Product '{product_name}' with part number '{part_number}' already exists.", 'error')
                continue  # Skip to the next row if the product already exists

            # Create the new product
            new_product = Product(
                productName=product_name,
                partNumber=part_number,
                productLabel=product_label,
                startingInventory=starting_inventory,
                inventoryIn=inventory_in,
                inventoryOut=inventory_out,
                inventoryOnHand=inventory_on_hand,
                minimumReq=minimum_req,
                image=url_for('static', filename='img/empty.jpg'),  # Placeholder image
                sale_price=sale_price,
                buy_price=buy_price
            )

            # Add the product to the session
            db.session.add(new_product)

        # Commit the changes to the database
        db.session.commit()

        flash("Products uploaded successfully.", 'success')
        return redirect("/productv")

    except Exception as e:
        db.session.rollback()  # Rollback in case of any error
        flash(f"Error: {str(e)}", 'error')
        return redirect("/")